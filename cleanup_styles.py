# cleanup_style.py
# Module to remove external links and excessive styles from an Excel workbook

from openpyxl import load_workbook
from cleanup import remove_excessive_styles
from cleanup import remove_excel_objects

# ------------------------------------------------------
# REMOVE EXTERNAL LINKS
# ------------------------------------------------------
def remove_external_links(wb):
    """
    Remove external links stored in:
      - external_links
      - defined names
      - formulas referring to other files
    """

    # Remove external link objects
    if hasattr(wb, "external_links"):
        wb.external_links = []
    if hasattr(wb, "_external_links"):
        wb._external_links = []

    # Remove external references in defined names
    to_delete = []

    for name in wb.defined_names:
        if name.attr_text and "[" in name.attr_text and "]" in name.attr_text:
            to_delete.append(name.name)

    for key in to_delete:
        del wb.defined_names[key]



# ------------------------------------------------------
# REMOVE PIVOT CACHES
# ------------------------------------------------------
def remove_pivot_caches(wb):
    if hasattr(wb, "_pivots"):
        wb._pivots = []

    if hasattr(wb, "_pivot_caches"):
        wb._pivot_caches = []


# ------------------------------------------------------
# MAIN PUBLIC FUNCTION (GUI & CLI CALL THIS)
# ------------------------------------------------------
def cleanup_styles_file(input_file, output_file=None):
    """
    Lightweight cleanup: remove excessive styles, drawings, and pivot caches.
    """
    if output_file is None:
        output_file = input_file.replace(".xlsx", "_STYLES_CLEANED.xlsx")

    wb = load_workbook(input_file, data_only=False)

    remove_excessive_styles(wb)
    remove_excel_objects(wb)
    remove_pivot_caches(wb)

    wb.save(output_file)
    return output_file


# ------------------------------------------------------
# FULL CLEANUP (EXTERNAL LINKS + STYLES + OBJECTS)
# Used by main.py --cleanup
# ------------------------------------------------------
def cleanup_excel_file(input_file, output_file=None):
    """
    Full cleanup pipeline:
        - external links
        - styles
        - drawings
        - pivot caches
    """
    if output_file is None:
        output_file = input_file.replace(".xlsx", "_CLEANED.xlsx")

    wb = load_workbook(input_file, data_only=False)

    remove_external_links(wb)
    remove_excessive_styles(wb)
    remove_excel_objects(wb)
    remove_pivot_caches(wb)

    wb.save(output_file)
    return output_file


# ------------------------------------------------------
# STANDALONE EXECUTION
# ------------------------------------------------------
if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Excel Style Cleanup Tool")
    parser.add_argument("input_file", help="Path to the Excel file")
    parser.add_argument("-o", "--output", help="Optional output file path")

    args = parser.parse_args()

    output = cleanup_styles_file(args.input_file, args.output)
    print(f"✔ Style cleanup complete! Saved to: {output}")
