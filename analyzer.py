"""analyzer.py
wb = openpyxl.load_workbook(path, data_only=False, read_only=False)
except Exception as e:
result['error'] = f'Failed to open workbook: {e}'
return result
"""
import zipfile
import re
from collections import Counter
from typing import Dict, Any, List
import openpyxl


VOLATILE_FUNCS = {'NOW', 'TODAY', 'INDIRECT', 'OFFSET', 'RAND', 'RANDBETWEEN'}
VOLATILE_REGEX = re.compile(r"\b(" + "|".join(VOLATILE_FUNCS) + r")\b", re.IGNORECASE)
FORMULA_REGEX = re.compile(r"^=.*")

def analyze_xlsx(path: str) -> Dict[str, Any]:
    result: Dict[str, Any] = {'path': path}


    # 1) Quick structure check
    try:
        with zipfile.ZipFile(path, 'r') as z:
            namelist = z.namelist()
            result['zip_entry_count'] = len(namelist)
            # media files
            media = [n for n in namelist if n.startswith('xl/media/')]
            result['media_count'] = len(media)
            # external links
            ext_links = [n for n in namelist if 'externalLinks' in n]
            result['external_links_count'] = len(ext_links)
    except zipfile.BadZipFile:
        result['error'] = 'BadZipFile: not a valid xlsx'
        return result


# 2) Workbook-level analysis with openpyxl
    try:
        wb = openpyxl.load_workbook(path, data_only=False, read_only=False)
    except Exception as e:
        result['error'] = f'Failed to open workbook: {e}'
        return result


    result['sheet_count'] = len(wb.sheetnames)
    sheets = {}
    total_cells = 0
    total_formulas = 0
    total_volatile = 0
    style_counter = Counter()
    merged_cells_count = 0


    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        info = {
            'max_row': ws.max_row,
            'max_column': ws.max_column,
            'formulas': 0,
            'volatile_formulas': 0,
            'merged_cells': len(list(ws.merged_cells.ranges)) if hasattr(ws, 'merged_cells') else 0,
            'hidden_rows': 0,
            'hidden_columns': 0,
            'unique_styles': 0,
        }


        # hidden rows/cols
        try:
            info['hidden_rows'] = sum(1 for r in ws.row_dimensions.values() if getattr(r, 'hidden', False))
            info['hidden_columns'] = sum(1 for c in ws.column_dimensions.values() if getattr(c, 'hidden', False))
        except Exception:
            pass


        styles_in_sheet = set()


    # iterate cells (careful: can be slow on huge files)
        for row in ws.iter_rows():
            for cell in row:
                total_cells += 1
                # style
                try:
                    if cell.has_style:
                        styles_in_sheet.add((cell.font, cell.fill, cell.border, cell.number_format, cell.alignment))
                except Exception:
                    pass
                # formula
                if cell.data_type == 'f' or (isinstance(cell.value, str) and FORMULA_REGEX.match(cell.value)):
                    info['formulas'] += 1
                    total_formulas += 1
                    val = cell.value if cell.value else ''
                    if VOLATILE_REGEX.search(str(val)):
                        info['volatile_formulas'] += 1
                        total_volatile += 1


        info['unique_styles'] = len(styles_in_sheet)
        info['merged_cells'] = info['merged_cells']
        merged_cells_count += info['merged_cells']
        sheets[sheetname] = info


    result.update({
        'total_cells_scanned_estimate': total_cells,
        'total_formulas': total_formulas,
        'total_volatile_formulas': total_volatile,
        'total_merged_cells': merged_cells_count,
        'sheets': sheets,
    })


    return result


if __name__ == '__main__':
    import sys, json
    if len(sys.argv) < 2:
        print('Usage: python analyzer.py <file.xlsx>')
        sys.exit(1)
    path = sys.argv[1]
    res = analyze_xlsx(path)
    print(json.dumps(res, indent=2))